import re

from django.shortcuts import render

from rest_framework.decorators import api_view, permission_classes, parser_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import HttpResponse

from openpyxl.cell import Cell
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl import Workbook
from openpyxl.styles import Font

from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required

from django.contrib.auth.models import User
from polls.models import UserProfile, Question, Poll, UserResponse

from rofr import keyconfig

from polls.utils import hasNumbers

# General USER Endpoints

@api_view(["POST"])
def register(request):
    """
        Register a new user and create their UserProfile
    """
    data = request.data
    
    try:

        # Email input check
        email = str(data['email_id'])
        if not re.match(r"(^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$)", email):
            return Response({'message':'Please enter a valid email address.'}, status=403)
        
        # Unique username check
        username = str(data["username"])
        try:
            user = User.objects.get(username=username)
            return Response({"message": "Username taken"}, status=412)
        except User.DoesNotExist:
            pass
        
        # Already registered email check
        if UserProfile.objects.filter(auth_user__email=email).exists():
            return Response({"message": "Email already registered"}, status=412)

        else:
            first_name = str(data['first_name'])
            last_name = data['last_name']
            password = str(data["password"])
            confirm_pass = str(data["confirm_pass"])
            if password != confirm_pass:
                return Response({"message": "Passwords don't match"}, status=412)

            if last_name == None:
                last_name = ''
            
            if (hasNumbers(first_name) or hasNumbers(last_name)):
                return Response({"message": "Invalid first/last name format"}, status=412)

            gender = str(data['gender']) # M/F/O
            valid_values = ["M", "F", "O"]
            if gender not in valid_values:
                return Response({"message": "Invalid gender value"}, status=412)

            age = int(data["age"])

            if not age > 0:
                return Response({"message": "Invalid age value"}, status=412)
            
            location = str(data["location"])
            if location == "":
                return Response({"message": "Please enter a location"}, status=412)
            
            income = int(data['income'])
            if income < 0:
                return Response({"message": "Invalid income value"}, status=412)
            
            try:
                # Create django User
                user = User()
                user.username = username
                user.email = email
                user.set_password(password)
                user.first_name = first_name
                user.last_name = last_name
                user.save()
                try:
                    # Create UserProfile
                    profile = UserProfile()
                    profile.auth_user = user
                    profile.gender = gender
                    profile.age = age
                    profile.income = income
                    profile.location = location
                    profile.save()
                    return Response({"message": "Registered!"}, status=201)
                except:
                    user.delete()
                    return Response({"message": "An error occurred. Please try again!"}, status=500)
            except:
                return Response({"message": "An error occurred. Please try again!"}, status=500)
    except KeyError as missing_data:
        return Response({'message':'Data is Missing: {}'.format(missing_data)}, status=400) 

@api_view(["GET"])
@permission_classes((IsAuthenticated,))
def available_polls(request):
    """
        Returns polls available to the user for attempting
    """       
    user = request.user

    polls = Poll.objects.all()
    payload = {
        "available_polls": [],
        "attempted_polls": []
    }
    for poll in polls:
        if poll in user.profile.attempted_polls.all():
            continue
        if (user.profile.age < poll.min_age or user.profile.income < poll.min_income):
            continue
        if (poll.max_age !=0 or poll.max_income != 0):
            if (user.profile.income > poll.max_income or user.profile.age > poll.max_age):
                continue
        questions = poll.questions.all()
        questions_data = []
        for question in questions:
            questions_data.append({
                "id": question.id,
                "title": question.title,
                "type": question.get_category_display(),
                "optional": question.optional
            })
        payload["available_polls"].append({
            "title": poll.title,
            "attempted_count": poll.attempted_count,
            "id": poll.id,
            "no_of_questions": poll.questions.all().count(),
            "questions": questions_data
        })
    
    for poll in user.profile.attempted_polls.all():
        payload["attempted_polls"].append({ # polls history of the user
            "title": poll.title
        })
        
    return Response(payload, status=200)

@api_view(["POST"])
@permission_classes((IsAuthenticated,))
def attempt_poll(request, poll_id):
    """
        Allows users to attempt a poll by ID
    """
    print("*** POLL ID: ***", poll_id)
    user = request.user
    profile = user.profile
    data = request.data

    if user.is_staff:
        return Response({"message": "Admin can't attempt a poll"}, status=412)

    try:
        try:
            poll = Poll.objects.get(id=poll_id)
            if poll in profile.attempted_polls.all():
                print("***** BLABLABABLA ******")
                return Response({"message": "You have already taken this poll"}, status=412)

            print("*** OUTSIDE FOR *****")
            question_responses = data["question_responses"]
            if len(question_responses) > poll.questions.all().count():
                return Response({"message": "Invalid number of question responses"}, 412)
        except Poll.DoesNotExist:
            return Response({"message": "Poll not found!"}, status=404)

        for question_response in question_responses:
            try:
                question_id = question_response["id"]
                try:
                    question = Question.objects.get(id=question_id)
                except Question.DoesNotExist:
                    return Response({"message": "Invalid question ID"}, status=404)    
                response_input = question_response["response_input"]

                print("***A QUESTION RESPONSE**")

                if (question.optional == False and response_input == ""):
                    poll_responses = poll.responses.all()
                    poll_responses.delete()
                    return Response({"message": "Invalid response to question with ID: {}".format(question.id)}, status=412)

                if (question.category == "S"): # if a Scale (1-5) type question
                    try:
                        response_input = int(response_input)
                        if response_input not in range(1, 6):
                            poll_responses = poll.responses.all()
                            poll_responses.delete()
                            return Response({"message": "Question response value must be between 1 and 5"}, status=412)
                    except ValueError:
                        print("***** VALUE ERROR ****")
                        poll_responses = poll.responses.all()
                        poll_responses.delete()
                        return Response({"message": "Invalid response to question with ID: {}".format(question.id)}, status=412)

                print("*** SAB SAHIII***")
                response_instance = UserResponse.objects.create(question=question, response=response_input, poll_taker=profile, poll=poll)
                print("*** RESPONSE INSTANCE SAVED ***")

            except KeyError as missing_data:
                return Response({'message':'Data is Missing: {}'.format(missing_data)}, status=400)

        user.profile.attempted_polls.add(poll)
        user.profile.save()
        poll.attempted_count += 1
        poll.save()
        return Response({"message": "Poll taken successfully!"}, status=200)

    except KeyError as missing_data:
        return Response({'message':'Data is Missing: {}'.format(missing_data)}, status=400) 


# Admin API Endpoints

@api_view(["POST"])
@permission_classes((IsAuthenticated,))
def create_poll(request):
    """
        Allows admins to create a new poll
    """       
    user = request.user
    if not user.is_staff:
        return Response({"message": "Unauthorized"}, status=401)
    data = request.data
    try:
        title = str(data["title"])
        questions = data["questions"]
        try:
            poll = Poll.objects.get(title=title)
            return Response({"message": "Poll with this title already exists"}, status=412)
        except Poll.DoesNotExist:
            poll = Poll.objects.create(title=title)
            try:
                if data["min_age"] != "":
                    poll.min_age = int(data["min_age"])
            except:
                pass    
            try:
                if data["max_age"] != "":
                    poll.max_age = int(data["max_age"])
            except:
                pass
            try:
                if data["min_income"] != "":
                    poll.min_income = int(data["min_income"])
            except:
                pass
            try:
                if data["max_income"] != "":
                    poll.max_income = int(data["max_income"])            
            except:
                pass
        for question in questions:
            valid_categories = ["MCSO", "MCMO", "OWNI", "S", "E"]
            try:
                question_title = str(question["question_title"])
                category = str(question["category"])
                if category not in valid_categories:
                    poll.delete()
                    return Response({"message": "Invalid question category"}, status=412)
                try:
                    optional = question["optional"]
                    valid_optionals = [True, False]
                    if optional not in valid_optionals:
                        poll.delete()
                        return Response({"message": "Invalid value for optional"}, status=412)
                except:
                    optional = False
                question = Question()
                question.title = question_title
                question.category = category
                question.optional = optional
                question.poll = poll
                question.save()
            except KeyError as missing_data:
                poll.delete()
                return Response({'message':'Data is Missing: {}'.format(missing_data)}, status=400) 
        return Response({"message": "Poll created!"}, status=201)
    except KeyError as missing_data:
        return Response({'message':'Data is Missing: {}'.format(missing_data)}, status=400) 

@api_view(["GET"])
@permission_classes((IsAuthenticated,))
def get_poll_responses(request, poll_id):
    """
        Returns responses for a particular poll
    """       

    if not request.user.is_staff:
        return Response({"message": "Unauthorized"}, status=401)

    try:
        poll = Poll.objects.get(id=poll_id)
    except Poll.DoesNotExist:
        return Response({"message": "Poll not found"}, status=404)
    
    poll_responses = poll.responses.all()
    if len(poll_responses) == 0:
        return Response({"message": "No one has taken this poll yet"}, status=200)

    payload = {
        "title": poll.title,
        "attempted_count": poll.attempted_count,
        "download_link": "localhost:8000/polls/responses/download/{}".format(poll_id),
        "responses": []
    }
    for response in poll_responses:
        payload["responses"].append({
            "full_name": response.poll_taker.auth_user.first_name + ' ' + response.poll_taker.auth_user.last_name,
            "response_input": response.response,
            "question_id": response.question.id,
            "question_category": response.question.get_category_display(),
            "queston_title": response.question.title
        })
            
    return Response(payload, status=200)

@api_view(["GET"])
@permission_classes((IsAuthenticated,))
def get_question_responses(request, q_id, poll_id):
    """
        Returns responses for a particular question of a particular poll
    """       

    if not request.user.is_staff:
        return Response({"message": "Unauthorized"}, status=401)
    
    try:
        poll = Poll.objects.get(id=poll_id)
        try:
            question = Question.objects.get(id=q_id)
        except Question.DoesNotExist:
            return Response({"message": "Question not found"}, status=404)
    except Poll.DoesNotExist:
        return Response({"message": "Poll not found"}, status=404)
        
    question_responses = UserResponse.objects.filter(poll=poll, question=question)

    if len(question_responses) == 0:
        return Response({"message": "No one has answered this question yet"}, status=200)

    payload = {
        "title": question.title,
        "optional": question.optional,
        "poll_title": question.poll.title,
        "category": question.get_category_display(),
        "attempted_count": len(question_responses),
        "download_link": "localhost:8000/polls/responses/download/{}/{}".format(poll_id, q_id),
        "responses": []
    }
    for response in question_responses:
        payload["responses"].append({
            "full_name": response.poll_taker.auth_user.first_name + ' ' + response.poll_taker.auth_user.last_name,
            "response_input": response.response,
        })
            
    return Response(payload, status=200)

@api_view(["GET"])
@permission_classes((IsAuthenticated,))
def aggregate_responses(request):
    """
        Returns the aggregate of the responses for the questions where applicable (Scale type questions)
    """
    if not request.user.is_staff:
        return Response({"message": "Unauthorized"}, status=401)
    
    scale_questions = Question.objects.filter(category="S")
    payload = {
        "aggregate_responses": []
    }
    for question in scale_questions:
        aggregate = 0
        for response in UserResponse.objects.filter(question=question):
            aggregate += int(response.response)
            payload["aggregate_responses"].append({
                "poll": question.poll.title,
                "question": question.title,
                "aggregate_response": aggregate
            })
    
    if len(payload["aggregate_responses"]) == 0:
        return Response({"message": "No Scale (1-5) type questions for providing aggregate responses"}, status=200)
    
    return Response(payload, status=200)


@staff_member_required(login_url="admin:login")
def poll_responses_excel(request, poll_id):
    """
        Receives a poll id and generates an excel sheet for the responses of that poll
    """
    poll = Poll.objects.get(id=poll_id)
    poll_responses = poll.responses.all()

    wb = Workbook()
    ws = wb.active
    title = "Responses for {}".format(poll.title)
    ws.title = title

    ws["A1"] = "Full Name"
    ws["B1"] = "Response"
    ws["C1"] = "Question Category"
    ws["D1"] = "Question"

    bold_font = Font(bold=True)

    for row in ws.iter_rows(min_row=1, max_col=4, max_row=1):
        for cell in row:
            cell.font = bold_font

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40
     
    row = 2

    for response in poll_responses:
        ws["A{}".format(row)] = response.poll_taker.auth_user.first_name + ' ' + response.poll_taker.auth_user.last_name
        ws["B{}".format(row)] = response.response
        ws["C{}".format(row)] = response.question.get_category_display()
        ws["D{}".format(row)] = response.question.title

        row += 1		

    response = HttpResponse(content=save_virtual_workbook(wb), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=Poll Responses.xlsx"
    return response

@staff_member_required(login_url="admin:login")
def question_responses_excel(request, poll_id, q_id):
    """
        Receives a poll id, question id and generates an excel sheet for the responses for that question of the particular poll
    """
    poll = Poll.objects.get(id=poll_id)
    question = Question.objects.get(id=q_id)
    question_responses = UserResponse.objects.filter(poll=poll, question=question)

    wb = Workbook()
    ws = wb.active
    ws.title = "Question Responses"

    ws["A1"] = "Full Name"
    ws["B1"] = "Question Title"
    ws["C1"] = "Poll"
    ws["D1"] = "Question Category"
    ws["E1"] = "Optional"
    ws["F1"] = "Response"


    bold_font = Font(bold=True)

    for row in ws.iter_rows(min_row=1, max_col=6, max_row=1):
        for cell in row:
            cell.font = bold_font

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 40

     
    row = 2

    for response in question_responses:
        optional = "No"
        if question.optional:
            optional = "Yes"
        
        ws["A{}".format(row)] = response.poll_taker.auth_user.first_name + ' ' + response.poll_taker.auth_user.last_name
        ws["B{}".format(row)] = question.title
        ws["C{}".format(row)] = poll.title
        ws["D{}".format(row)] = response.question.get_category_display()
        ws["E{}".format(row)] = optional
        ws["F{}".format(row)] = response.response

        row += 1		

    response = HttpResponse(content=save_virtual_workbook(wb), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response['Content-Disposition'] = "attachment; filename=Question Responses.xlsx"
    return response



    

